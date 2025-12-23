

#* ======================================================== *#
#*                     USEFUL SNIP-ITS                      *#
#* ======================================================== *#

#* GET CURRENT TIME
import time
time.time()

#* LAMBDA FUNCTIONS

#* DATACLASSES
from dataclasses import dataclass, field
#? SLOTS = MEMORY ENHANCEMENT / NO NEW ATTRIBUTES
#? KW_ONLY = FIELDS MUST BE ASSIGNED BY KEYWORD !LOCATION
@dataclass(slots=True, kw_only=True)
class Deer():
    deer: str
    eat: int
    chain: int
    
#* *ARGS, **KWARGS, AND * ---------------------------------  #
#? CAN BE PASSED KEYWORD ARGUMENTS THROUGH NESTED FUNCTION CALLS
def func1(d, e, c): pass
def func2(*, deer, eat, chain):
    return 
#TODO: ^^^^^^

#~ ======================================================== ~#
#~                    CLASS DEFINITION                      ~#
#~ ======================================================== ~#
    #& CONTEXT MANAGER 1/2
    def __enter__(self):
        if gvar.LOG_MSG:
            log.border()
            log.watchdog(' STARTING ')
        return self
    #& CONTEXT MANAGER 2/2
    def __exit__(self, exc_type, exc_val, exc_tb):
        if gvar.LOG_MSG:
            log.watchdog(' SHUTTING DOWN ')
            log.border()
        return False

#? ======================================================== ?#
#?                     UTIL FUNCTIONS                       ?#
#? ======================================================== ?#

#* EXTRACT BASE URL FOR ROBOTS.TXT CHECK ------------------  #
from urllib.parse import urlparse, urlunparse
def base_url(url):
    parsed_url = urlparse(url)
    return urlunparse((parsed_url.scheme, parsed_url.netloc, '', '', '', ''))

from util_classes import ColorLog
log = ColorLog('PYTHON EXAMPLES')




#? ======================================================== ?#
#?                      FILE OPERATIONS                     ?#
#? ======================================================== ?#

#* WRITE/OVERWRITE/READ FILE ------------------------------  #
file_data = 'NEW FILE CREATED'
file_path = 'DEEREATCHAIN/test.py'
# 'X'= ENFORCE NO OVERWRITE, 'W'= WRITE, 'R'= READ
with open(file_path, 'x', encoding='utf-8') as f:
                f.write(file_data)
                f.close()

#* CREATE FOLDER / TREE -----------------------------------  #
import os
folder_name = '/workspaces/DEC/DEEREATCHAIN'
if not os.path.exists(folder_name):
            os.makedirs(folder_name)

#* SEARCH FOR <FILE-NAME> STARTING IN ROOT FOLDER ---------  #
def cur_find_file(self, filename:str) -> str | None:
    root = self.root_path
    # RETURN ALL ITEMS LOWER THAN <ROOT> THAT MATCH <FILENAME>
    for file in root.rglob(filename):
        if file.is_file():
            return str(file) #!EXIT!
    return None  #!EXIT! FAILED TO FIND


#* [OPENPYXEL] - EXCEL FILE EDITING -----------------------  #
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
#? LOAD WORKBOOK / XLSM FILE
wb = load_workbook(file_path, keep_vba=True)
worksheets = wb.sheetnames
#? GET WORKSHEET    
ws = wb['worksheet']
#? UPDATE CELL IN WORKSHEET
ws.cell(row=1,column=1, value=0)
#? CREATE HEADER {COLUMN : INDEX} MAP 
def build_column_map(ws : Worksheet) -> dict:
        return {
            str(cell.value).strip().lower():cell.column
            for cell in ws[1]
            # IGNORE BLANKS
            if cell.value
        }


#^ ======================================================== ^#
#^                    SIMPLE MINI-MODULES                   ^#
#^ ======================================================== ^#

#* RETURN TEXT AS LOWERCASE WITH NO SPACES
def normalize_text(text : str):
    return ''.join(text.lower().split())




#? ======================================================== ?#
#?                    FREE SIMPLE GUI                       ?#
#? ======================================================== ?#
import FreeSimpleGUI as sg


sg.popup(
    args=*<1 or N object>,
    title = None,
    button_color = None,
    background_color = None,
    text_color = None,
    button_type = 0,
    auto_close = False,
    auto_close_duration = None,
    custom_text = (None, None),
    non_blocking = False,
    icon = None,
    line_width = None,
    font = None,
    no_titlebar = False,
    grab_anywhere = False,
    keep_on_top = None,
    location = (None, None),
    relative_location = (None, None),
    any_key_closes = False,
    image = None,
    modal = True,
    button_justification = None,
    drop_whitespace = True)
sg.popup('popup')  # Shows OK button
sg.popup_scrolled(args=*<1 or N object>,
    title = None,
    button_color = None,
    background_color = None,
    text_color = None,
    yes_no = False,
    no_buttons = False,
    button_justification = "l",
    auto_close = False,
    auto_close_duration = None,
    size = (None, None),
    location = (None, None),
    relative_location = (None, None),
    non_blocking = False,
    no_titlebar = False,
    grab_anywhere = False,
    keep_on_top = None,
    font = None,
    image = None,
    icon = None,
    modal = True,
    no_sizegrip = False)
sg.popup_get_text(message,
    title = None,
    default_text = "",
    password_char = "",
    size = (None, None),
    button_color = None,
    background_color = None,
    text_color = None,
    icon = None,
    font = None,
    no_titlebar = False,
    grab_anywhere = False,
    keep_on_top = None,
    location = (None, None),
    relative_location = (None, None),
    image = None,
    history = False,
    history_setting_filename = None,
    modal = True)
sg.popup_get_file(message,
    title = None,
    default_path = "",
    default_extension = "",
    save_as = False,
    multiple_files = False,
    file_types = (('ALL Files', '*.* *'),),
    no_window = False,
    size = (None, None),
    button_color = None,
    background_color = None,
    text_color = None,
    icon = None,
    font = None,
    no_titlebar = False,
    grab_anywhere = False,
    keep_on_top = None,
    location = (None, None),
    relative_location = (None, None),
    initial_folder = None,
    image = None,
    files_delimiter = ";",
    modal = True,
    history = False,
    show_hidden = True,
    history_setting_filename = None)
sg.popup_get_folder(message,
    title = None,
    default_path = "",
    no_window = False,
    size = (None, None),
    button_color = None,
    background_color = None,
    text_color = None,
    icon = None,
    font = None,
    no_titlebar = False,
    grab_anywhere = False,
    keep_on_top = None,
    location = (None, None),
    relative_location = (None, None),
    initial_folder = None,
    image = None,
    modal = True,
    history = False,
    history_setting_filename = None)
sg.popup_animated(image_source,
    message = None,
    background_color = None,
    text_color = None,
    font = None,
    no_titlebar = True,
    grab_anywhere = True,
    keep_on_top = True,
    location = (None, None),
    relative_location = (None, None),
    alpha_channel = None,
    time_between_frames = 0,
    transparent_color = None,
    title = "",
    icon = None,
    no_buffering = False)
sg.one_line_progress_meter(title,
    current_value,
    max_value,
    args=*<1 or N object>,
    key = "OK for 1 meter",
    orientation = "v",
    bar_color = (None, None),
    button_color = None,
    size = (20, 20),
    border_width = None,
    grab_anywhere = False,
    no_titlebar = False,
    keep_on_top = None,
    no_button = False)
sg.
sg.popup_ok('popup_ok')  # Shows OK button
sg.popup_yes_no('popup_yes_no')  # Shows Yes and No buttons
sg.popup_cancel('popup_cancel')  # Shows Cancelled button
sg.popup_ok_cancel('popup_ok_cancel')  # Shows OK and Cancel buttons
sg.popup_error('popup_error')  # Shows red error button
sg.popup_timed('popup_timed')  # Automatically closes
sg.popup_auto_close('popup_auto_close')  # Same as PopupTimed