
#* ======================================================== *#
 #*                    FILE DESCRIPTION                   
 #* EXCEL DATABASE MANIPULATION SCRIPT
 #*  - USES OPENPYXL TO READ/WRITE XLSM EXCEL FILES
 #*  - USES KEYWORDS.PY TO MAP VARIABLE NAMES TO EXCEL COLUMN NAMES
 #*  - FUNCTIONS:
 #*      - ensure_keywords_module() : Ensures keywords.py is loaded
 #*      - safe_get_sheet() : Safely get a worksheet by name
 #*      - build_column_map() : Build a map of column names to indexes
 #*      - build_keyvar_map() : Build a map of keyword variable names to column indexes
 #*      - update_first_match() : Update first row matching a value in a column
 #*      - update_all_match() : Update all rows matching a value in a column
 #*      - update_first_by_name() : Update first row matching a value in a named column
 #*  - DEPENDENCIES:
 #*      - openpyxl
 #*      - keywords.py
 #*
 #*
 #*
 #* ======================================================== *#


'''- STANDARD MODULE IMPORTS -'''
from pprint import pprint
import os, sys, importlib.util
# ^ END STANDARD MODULES ^ #

import os, sys
from pathlib import Path
# set project root to DEEREATCHAIN (two levels up from this file)
ROOT = str(Path(__file__).resolve().parents[2])
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)
ROOT = str(Path(__file__).resolve().parents[3])
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

'''- CUSTOM MODULE IMPORTS -'''
from system.utils.design_models import Controller, PowerSupply
from system.gen.settings import LOG_MSG
from john_scraper import JohnScraper
from system.utils.util_classes import *
from system.utils.data_models import Equipment, LEDProd, PSU, Ctrlr

'''- 3RD PARTY MODULE IMPORTS -'''
#* IMPORT BEAUTIFULSOUP FOR HTML PARSING
from bs4 import BeautifulSoup 
#* IMPORT OPENPYXL FOR EXCEL XLSX/XLSM HANDLING
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
# ^^^ 3RD PARTY MODULE IMPORTS ^^^ #


#* - PRINT CURRENT WORKING DIR + SYS.PATH - #
import sys, os
print("Current Working Dir:", os.getcwd())
print("Sys Path:", sys.path)

from system.gen.keywords import LED_KEYWORDS, PSU_KEYWORDS, CTRLR_KEYWORDS, GENERIC_KEYWORDS


#~ ======================================================== ~#
#~                    CLASS DEFINITION                      ~#
#~ ======================================================== ~#
log = ColorLog('JOHN_EQUI', level=1)
#* EQUIPMENT DATABASE MANAGER CLASS
class EquipDB:
    def __init__(self, wb_path:str):
        # TODO: ADD DYNAMICALLY CREATED DICTIONARY OF:
        # TODO:  - WORKBOOKS
        # TODO:  - WORKSHEETS
        # TODO:  - (COLUMN , INDEX MAP)
        # TODO:  - 
        self.wb = load_workbook(wb_path, keep_vba=True)
        if LOG_MSG and self.wb:
            log.success(f" Workbook '{wb_path}' loaded successfully.")
            log.info(f'Workbook Sheets: {self.wb.sheetnames}')
        else:
            log.error(f" Failed to load Workbook '{wb_path}'.")
        
        #* CHECK FOR REQUIRED SHEETS AND STORE REFERENCES
        self.led_py_db = self.find_sheet(self.wb, 'led_py_db')
        self.psu_py_db = self.find_sheet(self.wb, 'psu_py_db')
        self.ctrlr_py_db = self.find_sheet(self.wb, 'ctrlr_py_db')
        self.db_worksheets = [self.led_py_db,
                              self.psu_py_db,
                              self.ctrlr_py_db]
        if all(sheet is not None for sheet in self.db_worksheets):
            if LOG_MSG: log.success('All required database sheets found!')
        
        if self.led_py_db and self.psu_py_db and self.ctrlr_py_db:
            self.led_keyvar_map = self._build_keyvar_map(self.led_py_db, LED_KEYWORDS, GENERIC_KEYWORDS)
            self.psu_keyvar_map = self._build_keyvar_map(self.psu_py_db, PSU_KEYWORDS, GENERIC_KEYWORDS)
            self.ctrlr_keyvar_map = self._build_keyvar_map(self.ctrlr_py_db, CTRLR_KEYWORDS, GENERIC_KEYWORDS)
            
            
#? ======================================================== ?#
#?                    HELPER FUNCTIONS                      ?#
#? ======================================================== ?#
    
    #* SAFELY GET SHEETS
    def find_sheet(self, wb : Workbook, sheet_name : str):
        if sheet_name not in wb.sheetnames:
            log.error(f" Sheet '{sheet_name}' not found in workbook.")
            return None #!EXIT!#
        return wb[sheet_name] #!EXIT!#

    #* FINDS AND RETURNS A MAP OF LOWER CASE COLUMN {NAME : INDEX}
    def _build_column_map(self, ws : Worksheet):
        return {
            str(cell.value).strip().lower():cell.column
            for cell in ws[1]
            # IGNORE BLANKS
            if cell.value
        }

    #* BUILD AND RETURN A MAP OF MATCHED {KEYWORD VARIABLE : COLUMN INDEX}
    def _build_keyvar_map(self, ws : Worksheet, keyword_dict1 : dict, keyword_dict2: dict):
        col_map = self._build_column_map(ws)
        log.debug(f" Column Map for Sheet '{ws.title}': {col_map}")
        
        keyvar_map = {}
        found_all = True
        for key, header_name in keyword_dict1.items():
            header_name_lc = header_name.lower()
            if header_name_lc in col_map:
                keyvar_map[key.lower()] = col_map[header_name_lc]
            else:
                for k, h in keyword_dict2.items():
                    h_lc = h.lower()
                    if h_lc in col_map:
                        keyvar_map[k.lower()] = col_map[h_lc]
                        break
            if key.lower() not in keyvar_map:
                if LOG_MSG: log.warning(f" Header '{header_name}' not found in sheet '{ws.title}'")
                found_all = False
        if LOG_MSG and found_all: log.success(f" All Headers in Sheet: '{ws.title}' were found!")
        return keyvar_map #!EXIT!#


    #* QUERY THE DATABASE FOR _____ | SET FUZZY TO FALSE FOR EXACT MATCHES ONLY
    def _query_database(self, target, fuzzy = True):
        pass


    #* ADD / UPDATE TIMESTAMP FOR LATEST UPDATE
    def _timestamp_row(self):
        pass


    #* LOOK AT TIMESTAMPS AND RECCOMMEND UPDATES
    def _suggest_updates(self):
        pass


    #* CHECK CERTIFICATION STATUSES
    def _run_cert_checks(self):
        pass


    #* CHECKS FOR EXISTING URL IN DB
    def url_check(self, url: str):
        pass
    

    #* CHECKS FOR EXISTING NAME IN DB
    def name_check(self, name: str):
        pass



 
 #? ======================================================== ?#
 #?                   EXTERNAL FUNCTIONS                     ?#
 #? ======================================================== ?#
 #* LOADS A NEW DATABASE FILE FROM GIVEN PATH
    def load_new_xlsm(self, path):
        # TODO: ADD LAST KNOWN PATH VARIABLE
        pass



 #* UPDATE 1 CELL IN ALL ROWS THAT MATCH_COL VALUE = MATCH_VAL
    def update_all_match(self, ws : Worksheet, match_col, match_val, update_col, new_val):
        # HOLD COUNT OF ROWS CHANGED
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[match_col-1].value == match_val:
                ws.cell(row=row[match_col-1].row, column=update_col, value=new_val)
                count += 1

        log.info(f" Updated {count} rows; Where {match_col} = {match_val}")


    #* UPDATES ALL ROWS BY: UPDATE_COL_NAME WITH NEW_VALUE IF MATCH_VALUE IN MATCH_COL_NAME 
    def update_first_by_name(self, ws, key_map, match_col_name, match_value, update_col_name, new_value):
        # NORMALIZE CASE FOR INPUT COL NAMES
        match_col_name = match_col_name.lower()
        update_col_name = update_col_name.lower()

        # CATCH NO MATCHING COLUMN NAMES FOUND
        if match_col_name not in key_map or update_col_name not in key_map:
            log.warning(f" Column not found in key_map: '{match_col_name}' or '{update_col_name}'")
            return False #!EXIT!#
        
        # GET INDEXES OF COL NAMES
        match_col = key_map[match_col_name]
        update_col = key_map[update_col_name]
        
        # UPDATE FIRST ROW FOUND w/ MATCH_VALUE IN MATCH_COL_NAME | NEW_VALUE 
        for row in ws.iter_rows(min_row=2, values_only=False):
            cell = row[match_col-1]
            if str(cell.value).strip().lower() == str(match_value).strip().lower():
                ws.cell(row=cell.row, column=update_col, value=new_value)
                log.info(f" Updated Row {cell.row} : '{update_col_name}' to '{new_value}'")
                return True #!EXIT!#
        
        log.warning(f" Value '{match_value}' not found in column '{match_col_name}'")



    
    #* CHECK IF URL EXISTS IN DATABASE
    #* YES: SCRAPE AGAIN, CHECK FOR UPDATES
    #* NO: SCRAPE AND ADD NEW ENTRY
    def scrape_new_url(self, url:str):
        #url = 'https://www.environmentallights.com/19072-px-spi-v2.html'    # LED TAPE
        #url = 'https://www.environmentallights.com/17125-dmx-4-8a.html'     # CONTROLLER
        #url = 'https://www.environmentallights.com/17152-hlg-600h-24a.html' # POWER SUPPLY
        if True: #self.url_check(url):
            if LOG_MSG:
                log.warning(f'URL ALREADY EXISTS IN DATABASE: {url}')
                log.info(f'SCRAPING FOR UPDATES: {url}')
            with JohnScraper(url) as scraper:
                if scraper.data is not None: 
                    if LOG_MSG:
                        #log.success(f'SCRAPED DATA: {scraper.data}')
                        log.info(f'ADDING NEW EQUIPMENT TO DATABASE...')
                        pprint(scraper.new_equipment)
                    self.add_equipm(scraper.new_equipment) #!ADD TO DB
                    self.wb.save('DEEREATCHAIN/assets/PYDB.xlsm')
                else:
                    if LOG_MSG: log.error(f'FAILED TO SCRAPE URL: {url}')
                    return None #!EXIT!#

    #* BUILD A ROW LIST FROM EQUIPMENT OBJ FOR APPENDING TO SHEET
    #* RETURNS A FORMATTED/ORDERED LIST OF ROW-CELL-VALUES TO BE DIRECTLY APPENDED
    def build_row(self, equipm: Equipment, ws: Worksheet, key_map: dict) -> list:
        row = ['!'] *len(key_map)  # PRE-FILL LIST WITH PLACEHOLDERS
        for attr in equipm.__slots__:
            if attr not in key_map:
                continue #!SKIP UNMAPPED ATTRIBUTES!#
            if attr in GENERIC_KEYWORDS:
                col_index = key_map[attr]
                value = str(getattr(equipm, attr))
                log.debug(f' GEN ATTR: {attr} | COL INDEX: {col_index} | VALUE: {value}')
                # INSERT VALUE AT CORRECT INDEX (ADJUST FOR 0-BASED LIST)
                row.insert(col_index-1, value)
            elif attr in LED_KEYWORDS:
                col_index = key_map[attr]
                value = str(getattr(equipm, attr))
                log.debug(f' LED ATTR: {attr} | COL INDEX: {col_index} | VALUE: {value}')
                # INSERT VALUE AT CORRECT INDEX (ADJUST FOR 0-BASED LIST)
                row.insert(col_index-1, value)
            elif attr in PSU_KEYWORDS:
                col_index = key_map[attr]
                value = str(getattr(equipm, attr))
                log.debug(f' PSU ATTR: {attr} | COL INDEX: {col_index} | VALUE: {value}')
                # INSERT VALUE AT CORRECT INDEX (ADJUST FOR 0-BASED LIST)
                row.insert(col_index-1, value)
            elif attr in CTRLR_KEYWORDS:
                col_index = key_map[attr]
                value = str(getattr(equipm, attr))
                log.debug(f' CTRLR ATTR: {attr} | COL INDEX: {col_index} | VALUE: {value}')
                # INSERT VALUE AT CORRECT INDEX (ADJUST FOR 0-BASED LIST)
                row.insert(col_index-1, value)
            if LOG_MSG: log.debug(f'Built Row: {row}')
        return row

    #* ADD NEW EQUIPMENT OBJECT TO RESPECTIVE DATABASE (AUTO@POST-SCRAPE)
    def add_equipm(self, equipm) -> bool:
        if equipm is None:
            if LOG_MSG: log.error('ADD_EQUIPM() WAS GIVEN NONE!')
            return False #!EXIT!#
        
        #* DETERMINE EQUIPMENT TYPE AND APPEND TO CORRECT SHEET
        if isinstance(equipm, LEDProd):
            if LOG_MSG: log.info('ADDING NEW LED EQUIPMENT TO DATABASE...')
            if self.led_py_db: #~ CHECK SHEET EXISTS ~#
                row = self.build_row(equipm, self.led_py_db, self.led_keyvar_map)
                self.led_py_db.append(row) #~ APPEND NEW ROW ~#
                pprint(row)
        elif isinstance(equipm, PSU):
            if LOG_MSG: log.info('ADDING NEW POWER SUPPLY TO DATABASE...')
            if self.psu_py_db: #~ CHECK SHEET EXISTS ~#
                row = self.build_row(equipm, self.psu_py_db, self.psu_keyvar_map)
                self.psu_py_db.append(row) #~ APPEND NEW ROW ~#
        elif isinstance(equipm, Ctrlr):
            if LOG_MSG: log.info('ADDING NEW CONTROLLER TO DATABASE...')
            if self.ctrlr_py_db: #~ CHECK SHEET EXISTS ~#
                row = self.build_row(equipm, self.ctrlr_py_db, self.ctrlr_keyvar_map)
                self.ctrlr_py_db.append(row) #~ APPEND NEW ROW ~#
        return True #!EXIT!#


    #* REMOVE EQUIPMENT ENTRY FROM DATABASE
    def remove_equipm(self):
        pass
    
    
    #* SELECT EQUIPMENT ENTRY AND RETURNS EQUIPMENT OBJ.
    def pull_equipm(self):
        pass
    

    #* LOAD AND EDIT EQUIPMENT ENTRY
    def edit_equipm(self, equipm: Equipment):
        pass 
    


 #^ ======================================================== ^#
 #^                   TESTING / EXAMPLES                     ^#
 #^ ======================================================== ^#
 
def test():
    db_path = 'DEEREATCHAIN/assets/ENCL.xlsm'
    equip_db = EquipDB(db_path)
     
    # TEST SCRAPING AND ADDING NEW EQUIPMENT
    test_url = 'https://www.environmentallights.com/20914-plnw-pf-rgb30k-10m.html'
    equip_db.scrape_new_url(test_url)
    log.info('TEST COMPLETE.')
    
test()