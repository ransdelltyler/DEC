'''
/**

EXCEL DATABASE MANIPULATION SCRIPT
- USES OPENPYXL TO READ/WRITE XLSM EXCEL FILES
- USES KEYWORDS.PY TO MAP VARIABLE NAMES TO EXCEL COLUMN NAMES
- FUNCTIONS:
    - ensure_keywords_module() : Ensures keywords.py is loaded
    - safe_get_sheet() : Safely get a worksheet by name
    - build_column_map() : Build a map of column names to indexes
    - build_keyvar_map() : Build a map of keyword variable names to column indexes
    - update_first_match() : Update first row matching a value in a column
    - update_all_match() : Update all rows matching a value in a column
    - update_first_by_name() : Update first row matching a value in a named column
- DEPENDENCIES:
    - openpyxl
    - keywords.py

//////////////////////////////////////////////////////////////////////////////////////

'''
# 
#? ------ START IMPORT BLOCK ------ #

'''- STANDARD MODULE IMPORTS -'''
import os, sys, importlib.util
# ^ END STANDARD MODULES ^ #

'''- CUSTOM MODULE IMPORTS -'''
from util_classes import *


'''- 3RD PARTY MODULE IMPORTS -'''
#* IMPORT BEAUTIFULSOUP FOR HTML PARSING
from bs4 import BeautifulSoup 
#* IMPORT REQUESTS FOR HTTP REQUESTS
import requests
#* IMPORT RESPONSES FOR MOCKING REQUESTS IN TESTS
import responses
#* IMPORT PYPERCLIP FOR CLIPBOARD HANDLING
import pyperclip
#* IMPORT OPENPYXL FOR EXCEL XLSX/XLSM HANDLING
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
# ^^^ 3RD PARTY MODULE IMPORTS ^^^ #

#* - PRINT CURRENT WORKING DIR + SYS.PATH - #
import sys, os
print("Current Working Dir:", os.getcwd())
print("Sys Path:", sys.path)

#& KEYWORDS FILE PATH
KEYWORDS_PATH = r"keywords.py"

#& EXCEL DB COLUMN KEYWORDS
from keywords import (
    LED_KEYWORDS,
    PSU_KEYWORDS,
    CTRLR_KEYWORDS,
    GENERIC_KEYWORDS
)

log = ColorLog('main_log', level=logging.DEBUG)

#? ------ START FUNCTION BLOCK ------ #

#& ENSURES IT CAN ACTUALLY FIND THE KEYWORDS.PY MODULE
def ensure_keywords_module(path=KEYWORDS_PATH):
    # TRY TO FIND THE FILE
    if not os.path.exists(path):
        log.error(f" ERROR: Could not find keywords.py at: {path}")
        log.debug(" Make sure the path is correct and file is not renamed or moved.")
        return None

    module_name = "keywords"
    if module_name in sys.modules:
        log.info(" Reloading existing keywords module...")
        importlib.reload(sys.modules[module_name])
    else:
        # Add directory to sys.path if not already present
        base_dir = os.path.dirname(path)
        if base_dir not in sys.path:
            sys.path.append(base_dir)
        log.success(f" Added {base_dir} to sys.path")

    # ATTEMPT TO IMPORT
    try:
        import keywords
        log.success(" keywords.py successfully imported!")
        # Check contents
        for required in ["LED_KEYWORDS", "PSU_KEYWORDS", "CTRLR_KEYWORDS", "GENERIC_KEYWORDS"]:
            if not hasattr(keywords, required):
                log.error(f" Missing expected variable: {required}")
        return keywords
    except Exception as e:
        log.critical(f" Import failed: {e}")
        return None
# ^^^ ENSURE KEYWORDS MODULE IS PRESENT ^^^ #


#& RUN MODULE PRECHECKS 
def run_prechecks():
    # KEYWORDS MODULE
    keywords = ensure_keywords_module()
    if keywords is None:
        raise ImportError(" keywords.py could not be loaded. Stopping script.")
    
    # OTHER MODULES CAN BE CHECKED HERE

    # CHECK COMPLETE
    log.watchdog(" All prechecks completed successfully.")
# ^^^ RUN SCRAPER PRE-CHECKS ^^^ #


#& SAFELY LOAD WORKBOOK
def load_workbook_safe(path : str) -> Workbook | None:
    try:
        wb = load_workbook(path, keep_vba=True)
        log.success(f" Workbook '{path}' loaded successfully.")
        
        # PRINT CHECK - WORKBOOK SHEET NAMES | [NAME]
        log.info(f" Workbook Sheet Names : {wb.sheetnames}")

        return wb
    
    except Exception as e:
        log.critical(f" ERROR loading workbook '{path}': {e}")
        return None
# ^^^ SAFELY LOAD WORKBOOK ^^^ #


#& SAFELY GET SHEETS
def safe_get_sheet(wb : Workbook, sheet_name : str):
    all_found = True
    if sheet_name not in wb.sheetnames:
        all_found = False
        log.error(f" Sheet '{sheet_name}' not found in workbook.")
        return None
    if all_found : log.success(f" Found All Sheets")
    return wb[sheet_name]
# ^^^ SAFELY GET SHEET ^^^ #

#& FINDS AND RETURNS A MAP OF LOWER CASE COLUMN {NAME : INDEX}
def build_column_map(ws : Worksheet):
    return {
        str(cell.value).strip().lower():cell.column
        for cell in ws[1]
        # IGNORE BLANKS
        if cell.value
    }
# ^^^ BUILD COLUMN MAP ^^^ #


#& BUILD AND RETURN A MAP OF MATCHED {KEYWORD VARIABLE : COLUMN INDEX}
def build_keyvar_map(ws : Worksheet, keyword_dict : dict):
    col_map = build_column_map(ws)
    keyvar_map = {}
    all_found = True

    for key, header_name in keyword_dict.items():
        header_name_lc = header_name.lower()
        if header_name_lc in col_map:
            keyvar_map[key.lower()] = col_map[header_name_lc]
        else:
            all_found = False
            log.warning(f" Header '{header_name}' not found in sheet '{ws.title}'")
    log.success(f" All Headers in Sheet: '{ws.title}' were found!")
    return keyvar_map
# ^^^ BUILD KEY-VARIABLE MAP ^^^ #


#& Update 1 cell with a new value by checking another value.
def update_first_match( ws : Worksheet,
                        match_col,
                        match_val,
                        update_col,
                        new_val) -> bool:

    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[match_col-1]
        if cell.value == match_val:
            ws.cell(row=cell.row, column=update_col, value=new_val)
            log.info(f" Updated row {cell.row} : set column {update_col} to {new_val}")
            return True
    log.warning(f" Value '{match_val}' not found in column {match_col}")
    return False
# ^^^ UPDATE FIRST FOUND MATCH ^^^ #

#& UPDATE 1 CELL IN ALL ROWS THAT MATCH_COL VALUE = MATCH_VAL
def update_all_match(ws : Worksheet, match_col, match_val, update_col, new_val):
    # HOLD COUNT OF ROWS CHANGED
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=False):
        if row[match_col-1].value == match_val:
            ws.cell(row=row[match_col-1].row, column=update_col, value=new_val)
            count += 1

    log.info(f" Updated {count} rows; Where {match_col} = {match_val}")

# ^^^ UPDATE ALL FOUND MATCHES ^^^ #

#& UPDATES ALL ROWS BY: UPDATE_COL_NAME WITH NEW_VALUE IF MATCH_VALUE IN MATCH_COL_NAME 
def update_first_by_name(ws, key_map, match_col_name, match_value, update_col_name, new_value):
    # NORMALIZE CASE FOR INPUT COL NAMES
    match_col_name = match_col_name.lower()
    update_col_name = update_col_name.lower()

    # CATCH NO MATCHING COLUMN NAMES FOUND
    if match_col_name not in key_map or update_col_name not in key_map:
        log.warning(f" Column not found in key_map: '{match_col_name}' or '{update_col_name}'")
        return False
    
    # GET INDEXES OF COL NAMES
    match_col = key_map[match_col_name]
    update_col = key_map[update_col_name]
    
    # UPDATE FIRST ROW FOUND w/ MATCH_VALUE IN MATCH_COL_NAME | NEW_VALUE 
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[match_col-1]
        if str(cell.value).strip().lower() == str(match_value).strip().lower():
            ws.cell(row=cell.row, column=update_col, value=new_value)
            log.info(f" Updated Row {cell.row} : '{update_col_name}' to '{new_value}'")
            return True
    
    log.warning(f" Value '{match_value}' not found in column '{match_col_name}'")

# ^^^ UPDATE ALL ROWS WITH MATCHING VALUE IN SELECTED COLUMN ^^^ #


# ^^^ END FUNCTION BLOCK ^^^ #


#? ------ START MAIN BLOCK ------ #


# EXCEL DATABASE SETUP W/ OPENPYXL
wb_path = r'ENCL.xlsm'

# LOAD DATABASE WORKBOOK XLSM
wb = load_workbook_safe(wb_path)
if wb is None:
    log.critical(" Could not load workbook. Exiting script.")
    sys.exit(1)

#* - GET DATABASE WORKSHEETS - #
# INIT AND SET TO EMPTY IN CASE SHEET IS NOT FOUND
key_map_led = key_map_psu = key_map_ctrl = key_map_gen = {}

# LED TAPE
ws_led = safe_get_sheet(wb,'LED Tape Data')
if ws_led: key_map_led = build_keyvar_map(ws_led, LED_KEYWORDS)

# PSUS
ws_psu = safe_get_sheet(wb,'PSU Data')
if ws_psu: key_map_psu = build_keyvar_map(ws_psu, PSU_KEYWORDS)

# CONTROLLERS
ws_ctrl = safe_get_sheet(wb, 'CTRLR Data')
if ws_ctrl: key_map_ctrl = build_keyvar_map(ws_ctrl, CTRLR_KEYWORDS)

# GENERIC
ws_gen = safe_get_sheet(wb, 'GEN Data')
if ws_gen: key_map_gen = build_keyvar_map(ws_gen, GENERIC_KEYWORDS)

#! PRINT CHECK - DICTIONARIES | VARKEY : COL_INDEX
log.debug(f" [LED] Key Column Map : {key_map_led}")
log.debug(f" [PSU] Key Column Map : {key_map_psu}")
log.debug(f" [CTRLR] Key Column Map : {key_map_ctrl}")
log.debug(f" [GENERIC] Key Column Map : {key_map_gen}")

# WRAP MAPS INTO ONE OBJECT | SHEET_MAPS
sheet_maps = {
    'LED Tape Data' : key_map_led,
    'PSU Data' : key_map_psu,
    'CTRLR Data' : key_map_ctrl,
    'GENERIC Data' : key_map_gen
}

### ^ EXCEL DATABASE SETUP W/ OPENPYXL ^ ###


# SAVE THE FILE TO NEW WORKBOOK
new_db_path = r'ENCL_edited.xlsm'
wb.save(new_db_path)
print(' Saved as : ' + new_db_path)


# ------ END MAIN BLOCK ------ #


# ------ STANDARDS + CONVENTIONS BLOCK ------ #

